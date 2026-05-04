#!/usr/bin/env python3
"""Gera o template MASTER_RISK_REGISTER.xlsx para a diligencia da Casa de Criadores.

O arquivo final fica em ./templates/MASTER_RISK_REGISTER.xlsx, com duas abas:
- "Controle": linhas pre-preenchidas Tier 1, validacoes de dados, filtro automatico,
  primeira linha congelada e larguras ajustadas.
- "Listas": valores permitidos para Status, Prioridade, Red flag, Tipo de risco e Decisao.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


COLUMNS: list[tuple[str, int]] = [
    ("ID", 10),
    ("Documento solicitado", 48),
    ("Pasta destino", 38),
    ("Responsável", 16),
    ("Prioridade", 12),
    ("Status", 14),
    ("Data solicitação", 16),
    ("Data recebimento", 16),
    ("Red flag", 12),
    ("Tipo de risco", 18),
    ("Observação", 40),
    ("Impacto jurídico", 18),
    ("Impacto financeiro", 18),
    ("Impacto operacional", 20),
    ("Decisão", 22),
]

LISTS: dict[str, list[str]] = {
    "Status": [
        "Pendente",
        "Solicitado",
        "Recebido",
        "Incompleto",
        "Em análise",
        "Validado",
        "Não existe",
        "Red flag",
    ],
    "Prioridade": ["Tier 1", "Tier 2", "Tier 3"],
    "Red flag": ["Sim", "Não", "A verificar"],
    "Tipo de risco": [
        "Jurídico",
        "Financeiro",
        "Operacional",
        "Reputacional",
        "Fiscal",
        "Societário",
        "Marca/IP",
        "Editais/Compliance",
    ],
    "Decisão": [
        "Aguardar",
        "Cobrar responsável",
        "Escalar para jurídico",
        "Escalar para financeiro",
        "Red flag crítica",
        "Validado",
        "Não aplicável",
    ],
}

TIER_1_ROWS: list[dict[str, str]] = [
    {
        "ID": "CDC-001",
        "Documento solicitado": "Mapa de todos os CNPJs envolvidos",
        "Pasta destino": "01_ESTRUTURA_SOCIETARIA_E_JURIDICA",
        "Responsável": "André",
        "Tipo de risco": "Societário",
    },
    {
        "ID": "CDC-002",
        "Documento solicitado": "Contrato social / estatuto das entidades",
        "Pasta destino": "01_ESTRUTURA_SOCIETARIA_E_JURIDICA",
        "Responsável": "Flavio",
        "Tipo de risco": "Societário",
    },
    {
        "ID": "CDC-003",
        "Documento solicitado": "Documento de titularidade da marca Casa de Criadores",
        "Pasta destino": "02_MARCA_E_PROPRIEDADE_INTELECTUAL",
        "Responsável": "André",
        "Tipo de risco": "Marca/IP",
    },
    {
        "ID": "CDC-004",
        "Documento solicitado": "Lista de processos judiciais e passivos",
        "Pasta destino": "03_PASSIVOS_CONTENCIOSO",
        "Responsável": "Flavio",
        "Tipo de risco": "Jurídico",
    },
    {
        "ID": "CDC-005",
        "Documento solicitado": "Protestos ativos",
        "Pasta destino": "03_PASSIVOS_CONTENCIOSO",
        "Responsável": "Flavio",
        "Tipo de risco": "Jurídico",
    },
    {
        "ID": "CDC-006",
        "Documento solicitado": "Bloqueios bancários ou judiciais",
        "Pasta destino": "03_PASSIVOS_CONTENCIOSO",
        "Responsável": "Flavio",
        "Tipo de risco": "Jurídico",
    },
    {
        "ID": "CDC-007",
        "Documento solicitado": "Contratos de patrocínio ativos",
        "Pasta destino": "04_ATIVOS_E_RECEBIVEIS",
        "Responsável": "Alzi",
        "Tipo de risco": "Operacional",
    },
    {
        "ID": "CDC-008",
        "Documento solicitado": "Cronograma detalhado dos recebíveis previstos",
        "Pasta destino": "04_ATIVOS_E_RECEBIVEIS",
        "Responsável": "Contador",
        "Tipo de risco": "Financeiro",
    },
]

DEFAULT_OUTPUT = Path("templates/MASTER_RISK_REGISTER.xlsx")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)


def build_lists_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Listas")
    for col_idx, (name, values) in enumerate(LISTS.items(), start=1):
        letter = get_column_letter(col_idx)
        ws.cell(row=1, column=col_idx, value=name).font = HEADER_FONT
        ws.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws.cell(row=1, column=col_idx).alignment = HEADER_ALIGN
        for row_idx, value in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)
        ws.column_dimensions[letter].width = max(len(name), max(len(v) for v in values)) + 2

        last_row = 1 + len(values)
        defined_name = f"lista_{name.replace(' ', '_').replace('/', '_')}"
        wb.defined_names[defined_name] = DefinedName(
            name=defined_name,
            attr_text=f"Listas!${letter}$2:${letter}${last_row}",
        )


def column_index(header: str) -> int:
    return next(i for i, (name, _) in enumerate(COLUMNS, start=1) if name == header)


def build_controle_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Controle", 0)

    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    for row_offset, row_data in enumerate(TIER_1_ROWS, start=2):
        defaults = {
            "Prioridade": "Tier 1",
            "Status": "Pendente",
            "Red flag": "A verificar",
            "Decisão": "Aguardar",
        }
        merged = {**defaults, **row_data}
        for header, _ in COLUMNS:
            value = merged.get(header, "")
            cell = ws.cell(row=row_offset, column=column_index(header), value=value)
            cell.alignment = DATA_ALIGN

    apply_validations(ws)


def apply_validations(ws) -> None:
    bindings = [
        ("Status", "lista_Status"),
        ("Prioridade", "lista_Prioridade"),
        ("Red flag", "lista_Red_flag"),
        ("Tipo de risco", "lista_Tipo_de_risco"),
        ("Decisão", "lista_Decisão"),
    ]
    last_row = 1000
    for header, list_name in bindings:
        col_letter = get_column_letter(column_index(header))
        dv = DataValidation(
            type="list",
            formula1=f"={list_name}",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Valor inválido",
            error=f"Selecione um valor da lista '{header}'.",
        )
        dv.add(f"{col_letter}2:{col_letter}{last_row}")
        ws.add_data_validation(dv)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Caminho do arquivo Excel a gerar. Padrão: {DEFAULT_OUTPUT}",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    output_path: Path = args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    build_controle_sheet(wb)
    build_lists_sheet(wb)
    wb.active = wb["Controle"]

    wb.save(output_path)
    print(f"Gerado: {output_path.resolve()}")
    print(f"  - {len(TIER_1_ROWS)} linhas Tier 1 pré-preenchidas")
    print(f"  - {len(COLUMNS)} colunas")
    print(f"  - {len(LISTS)} listas de validação na aba 'Listas'")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
